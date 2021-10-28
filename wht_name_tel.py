import PySimpleGUI as sg
from openpyxl.cell import cell
from selenium import webdriver
from selenium.webdriver.common import keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as CondicaoEsperada
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
import pyautogui 
import pyperclip
import webbrowser
import threading
import urllib
import time
import os 
import sys
import socket
import random
import sqlalchemy
from sqlalchemy import create_engine, Column, Integer, String
from sqlalchemy.ext.declarative import declarative_base
from sqlalchemy.orm import sessionmaker
from openpyxl import load_workbook
import openpyxl
from datetime import datetime

Base = declarative_base()

class Contacts_whatsapp(Base):
    __tablename__ = 'contacts_whatsapp'
    contact_id = Column(Integer, primary_key = True)
    name_contact = Column(String)
    number = Column(String)

#criando banco de dados SQLite3
#sqlite:/// é para o bando de dados ser criado na pasta atual
engine=create_engine('sqlite:///contacts_whatsapp.db', echo = True)
Base.metadata.create_all(bind=engine)
#criando uma conexao
Conexao = sessionmaker(bind=engine)
conexao = Conexao()
#aqui estou definindo que a varial conexao é do tipo sessionmaker
conexao: sessionmaker

class Whatsapp_web:
    #funcao janela principal
    #larg x alt
    def window_main(self):
        sg.theme('Reddit')

        menu_tela_inicial=[
            ['Configurações',['Meus contatos']],
            ['Sobre',['Suporte']]
        ]
        column_connected=[
            [sg.Text('Conectado',visible=True, justification='left')]
        ]
        # um column dentro de column divide a column
        column_1=[
            [sg.Multiline('Digite aqui a mensagem que deseja enviar...',size=(32,14), key='campo_message',justification='left', background_color='white')],
            [sg.Input('aqui fica o caminho da imagem...',key='Input_imagem', enable_events=True,size=(34,1))],
            [sg.Checkbox(' Anexar imagem', key='anexar', enable_events=True), sg.FileBrowse('selecionar imagem', key='select_pictures', target='Input_imagem',file_types=(('Arquivos Imagens','*.png'),),disabled=True)],
            [sg.Button('Conectar', key='btn_conectar', size=(9,1)),sg.Button('Pausar',size=(8,1),disabled=True),sg.Button('Enviar',size=(8,1),disabled=True)],
            [sg.Column(column_connected, justification='left')]
        ]
        column_2=[
            [sg.Multiline(visible=True, key='selection_contacts',size=(50,18))],
            [sg.Input('caminho do arquivo', key='Input_do_caminho',size=(52,1),)],
            [sg.Button('Add',key='add_btn',disabled=False),sg.FileBrowse('selecionar arquivo', key="-IN-", target='Input_do_caminho',file_types=(('Arquivos do excel','*'),))]
        ]
        layout_w_main=[
            [sg.Menu(menu_tela_inicial)],
            [sg.Text('Enviador de Whatsapp')],
            [sg.Column(column_1,element_justification='right'),sg.Column(column_2,element_justification='right')],
        ]
        return sg.Window('Enviador de whatsapp',layout_w_main,finalize=True, element_justification='c', size=(700,408))
    pass

    def window_del(self):
        layout_w_del =[
            [sg.Text('Deseja realmente apagar todos os contatos da lista atual?')],
            [sg.Button('  Sim  ', key='-sim-'), sg.Button('  Não  ', key='-nao-')]
        ]
        return sg.Window('Enviador de whatsapp', layout_w_del, finalize=True, disable_close=True, element_justification='c')
    pass
        
    def Iniciar(self):
        print(os.path.dirname(os.path.realpath("chrome.exe")))
        window_1,window_2=self.window_main(),None
        for linha in conexao.query(Contacts_whatsapp).all():
            window_1['selection_contacts'].update(f'{linha.name_contact} | {linha.number}\n', append=True)
        while True:
            #window,event, values é um padrao de reconhecimento que o python usa
            window,event,values=sg.read_all_windows()
            if window == window_1 and event in (sg.WIN_CLOSED,None):
                conexao.commit()
                window.close()
                sys.exit()
            if window == window_1 and event == 'Meus contatos':
                window_2 = self.window_del()
                window_1.disable() 
            if window == window_2 and event == '-sim-':
                a_ser_excluido= conexao.query(Contacts_whatsapp).all()
                for linha in a_ser_excluido:
                    conexao.delete(linha)
                pass
                conexao.commit()
                window_1['Input_do_caminho'].update('caminho do arquivo')
                window_1['selection_contacts'].update('seus contatos para envio')
                window_1['Enviar'].update(disabled=True)
                window_1.enable()
                window_2.close()
            if window == window_2 and event =='-nao-':
                window_1.enable()
                window_2.close()

            if window == window_1 and event == 'btn_conectar':
                window_1['Enviar'].update(disabled=False)
                dir_absolutle = os.path.dirname(os.path.realpath(__file__))
                class Navegar:
                    def __init__(self):
                        #init declara o driver para ser acessado no self, acessar o programa todo
                        chrome_options = Options()
                        #caminho do chromium
                        #try:
                        #    chrome_options.binary_location = dir_absolutle + '\\chrome-win' + '\\chrome.exe'
                        #except:
                        #    chrome_options.binary_location = os.getcwd() + '\\chrome-win' + '\\chrome.exe'
                        #chrome_options.add_argument("--headless")  
                        #chrome_options.add_argument("--disable-infobars")
                        #chrome_options.add_argument("--disable-extensions")
                        #chrome_options.add_argument("disable-media-stream")
                        #chrome_options.add_argument("allow-file-access-from-files")
                        #usar som ou video falso
                        #chrome_options.add_argument("use-fake-device-for-media-stream")
                        #chrome_options.add_argument("use-fake-ui-for-media-stream")
                           # Pass the argument 1 to allow and 2 to block
                        chrome_options.add_experimental_option("prefs",{\
                            "hardware.audio_capture_enabled": False,
                            "hardware.video_capture_enabled": False,
                            "hardware.audio_capture_blocked_urls":["https://web.whatsapp.com/"],
                            "hardware.video_capture_blocked_urls":["https://web.whatsapp.com/"]
                        })
                        #"profile.default_content_setting_values.media_stream_mic" : 2}
                        #chrome_options.add_argument("--user-data-dir=chrome-data")
                        chrome_options.add_argument('--ignore-certificate-errors-spki-list')
                        chrome_options.add_argument('--ignore-certificate-errors')
                        chrome_options.add_argument('--ignore-ssl-errors')
                        chrome_options.add_argument('--lang=pt-BR')
                        chrome_options.add_argument('--disable-notifications')
                        chrome_options.add_argument('--disable-gpu')
                        args = ["hide_console", ] 
                        try:
                            caminho_chromedriver = dir_absolutle + '\\chromedriver.exe'
                        except:
                            caminho_chromedriver = os.getcwd() + '\\chromedriver.exe'
                        caminho_chromedriver = r'C:\Users\Daniel pc\Desktop\whatsappcdb\whatsapp_name_tel\chromedriver.exe'
                        self.driver = webdriver.Chrome(executable_path=caminho_chromedriver,options=chrome_options, service_args=args)
                        self.wait = WebDriverWait( #aqui a variavel self.wait está recebendo o webdriverwait com as caracteristicas
                        driver=self.driver,
                        timeout= 10,
                        poll_frequency=6
                                        )
                    pass
                    def acessa_site(self):
                        def is_connected():
                            try:
                                # connect to the host -- tells us if the host is actually
                                # reachable
                                socket.create_connection(("www.google.com", 80))
                                return True
                            except :
                                is_connected()
                        pass
                        self.driver.get("https://web.whatsapp.com/")
                    pass  

                    def send_message_for_numb(self):
                        Conexao = sessionmaker(bind=engine)
                        conexao = Conexao()
                        #aqui estou definindo que a varial conexao é do tipo sessionmaker
                        conexao: sessionmaker
                        ## Obter tudo registrado em uma tabela
                        for linha in conexao.query(Contacts_whatsapp).all():
                            time.sleep(5)
                            textao =values['campo_message']
                            time.sleep(3)
                            try:
                                #aqui o programa trocará fullname pelo nome completo da pessoa da lista
                                for fullname in textao:
                                    textao=textao.replace('fullname',linha.name_contact)   
                                pass
                            except:
                                print('falha full_name')
                                pass
                            time.sleep(3)
                            try:
                                #aqui o programa trocará firstname pelo primeiro nome do cliente.
                                for firstname in textao:
                                    textao=textao.replace('firstname',linha.name_contact.split()[0])
                            except:
                                print('falha first_name')
                                pass
                            time.sleep(3)
                            try:
                                #aqui onde você escrever xsaudacaox será substituído por Bom dia e Boa tarde
                                for xsaudacaox in textao:
                                    hora_saudacao ='12:00'
                                    horario_atual = datetime.now()
                                    hora = int(hora_saudacao.split(':')[0])
                                    if horario_atual.hour >= hora:
                                        textao=textao.replace('xsaudacaox', 'Boa tarde')
                                    else:
                                        textao=textao.replace('xsaudacaox', 'Bom dia')
                                    pass     
                                time.sleep(3)
                            except:
                                print('saudacao invalida')
                                pass
                            textoo=textao# matriz 
                            self.driver.implicitly_wait(30)
                            print(str(textoo[0]))
                            textoo=urllib.parse.quote(textoo)
                            time.sleep(3)
                            try:
                                link_gerado = (f"https://web.whatsapp.com/send?phone={linha.number}&text={textoo}")
                                self.driver.get(link_gerado)
                                self.driver.implicitly_wait(60)
                                time.sleep(3)
                                #self.digite_como_uma_pessoa(textoo, area_texto) 
                                try:
                                    button_ok=self.driver.find_element_by_xpath('//div[text()="OK"]')
                                    button_ok.click()
                                except:
                                    if values['anexar'] == True:
                                        self.send_pictures()
                                    else:
                                        try:
                                            area_texto =self.driver.find_element_by_xpath('//*[@id="main"]/footer/div[1]/div[2]/div/div[1]/div')
                                            area_texto.send_keys(Keys.ENTER)
                                            time.sleep(3)
                                        except:
                                            pass
                                    pass  
                            except:
                                print('algo deu errado, talvez url não encontrada')
                                pass
                            conexao.delete(linha)
                            time.sleep(2)
                        pass
                    pass
                    def send_pictures(self):
                        button_anexar=self.wait.until(CondicaoEsperada.element_to_be_clickable((By.XPATH,'//*[@id="main"]/footer/div[1]/div[1]/div[2]/div/div/span')))
                        button_anexar.click()
                        time.sleep(3)
                        attach = self.driver.find_element_by_xpath('//input[@accept="image/*,video/mp4,video/3gpp,video/quicktime"]')
                        attach.send_keys(str(values['Input_imagem']).replace('/','\\'))
                        time.sleep(2)
                        #aqui para baixo envia a mensagem
                        button_send_p=self.wait.until(CondicaoEsperada.element_to_be_clickable((By.XPATH,'//span[@data-icon="send"]')))
                        button_send_p.click()
                        time.sleep(3)
                    pass
                navegador = Navegar()
                #navegador.acessa_site()  
                threading_conectar=threading.Thread(target=navegador.acessa_site,args=(), daemon=True)
                threading_conectar.start()
            if window == window_1 and event =='Enviar':
                window_1['Pausar'].update(disabled=False)
                try:
                   threading_conectar.join()
                   time.sleep(3)
                except:
                    pass
                try:
                    window_1['Enviar'].update(disabled=True)
                    threading_send_message_for_numb=threading.Thread(target=navegador.send_message_for_numb,args=(), daemon=True)
                    threading_send_message_for_numb.start()
                except:
                    sg.Popup('Conecte o whatsapp primeiro antes de continuar!')
                pass
            if window == window_1 and event =='Pausar':
                window_1['Pausar'].update('Pausado')
                threading_send_message_for_numb.join()

            if window == window_1 and event == 'add_btn':
                window_1['Input_do_caminho'].update(disabled=True)
                window_1['btn_conectar'].update(disabled=False)
                #window_1['add_btn'].update(disabled=True)
                window_1['Enviar'].update(disabled=False)
                try:
                    caminho_arquivo = values['Input_do_caminho']
                    arquivo_excel_existente = load_workbook(caminho_arquivo)
                    window_1['selection_contacts'].update('')
                    planilha1= arquivo_excel_existente.active
                    max_linha = planilha1.max_row
                    for i in range(1, max_linha + 1):
                        linha=f'{planilha1.cell(column=1, row= 1+i).value} | {planilha1.cell(column=3, row= 1+i).value}\n'
                        #aqui adiciona dados no multiline
                        window_1['selection_contacts'].update(linha, append=True)
                        window_1['selection_contacts'].update(disabled=True)
                        novo_contact = Contacts_whatsapp()
                        novo_contact.number = planilha1.cell(column=3, row= 1+i).value
                        novo_contact.name_contact= planilha1.cell(column=1, row= 1+i).value
                        if novo_contact.number == None and novo_contact.name_contact == None:
                            break
                        else:
                            pass
                        #add um novo registro
                        conexao.add(novo_contact)
                        #Salvar a info no banco de dados
                        conexao.commit()
                    pass
                except FileNotFoundError:
                    sg.Popup('Selecione o arquivo com os contatos')
                    window_1['add_btn'].update(disabled=False)

            if window == window_1 and event == 'Suporte':
                sg.Popup('Suporte (32)9.8899-9920')

            if window == window_1 and values['anexar'] == True:
                window_1['select_pictures'].update(disabled=False)
                window_1['Input_imagem'].update(disabled=True)
            else:
                window_1['select_pictures'].update(disabled=True)
                window_1['Input_imagem'].update('')
                pass
        
        window_1.close()
    pass
whats=Whatsapp_web()
whats.Iniciar()
